import os
import glob

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

def addAddressee():
    file_paths = glob.glob(os.path.join(os.getcwd(), 'AnnotatedTextData', '*.xlsx'))
    for file_path in file_paths:
        print("Processing file:", file_path)

        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active

        worksheet['G1'] = 'addressee'
        worksheet['G1'].fill = PatternFill(start_color="808080", fill_type="solid")

        for row in worksheet.iter_rows(min_row=2):
            if row[3].value == 1:
                row[6].value = 2
                row[6].fill = PatternFill(start_color="CCCCCC", fill_type="solid")
            else:
                row[6].value = 1

                if row[3].value == 2:
                    row[6].fill = PatternFill(start_color="FFFFFF", fill_type="solid")
                else:
                    row[6].fill = PatternFill(start_color="808080", fill_type="solid")

        workbook.save(file_path)


def compileAnnotation():
    """
        Compiles annotations from multiple Excel files in a folder and saves the result in a new Excel file.

        The file name format is "participant_simulation_civilian_transcription_state.xlsx".

        Each input Excel file is expected to have the following columns:
        - Column 4: Speaker ID (1 or 2)
        - Column 5: Annotation for speaker ID 2
        - Column 6: Annotation for speaker ID 1

        Note: This function requires the openpyxl and pandas libraries to be installed.
        """

    # Fetches list(str) of directories to Excel files
    file_paths = glob.glob(os.path.join(os.getcwd(), 'AnnotatedTextData', '*.xlsx'))
    output_file = os.path.join(os.getcwd(), 'compiled_annotations.xlsx')

    compiled_annotations = pd.DataFrame(columns=["participant", "simulation", "civilian"])

    # Retrieves annotations and text from each file and generates Excel sheet
    for file_path in file_paths:
        print("Processing file:", file_path)

        # Extracts data from filename and creates headers
        filename = os.path.basename(file_path)
        participant, simulation, civilian, transcription, state = filename[:-5].split('_')
        new_row = {"participant": int(participant), "simulation": int(simulation), "civilian": civilian}

        df = pd.read_excel(file_path)

        turn = 1
        row = 0

        # If first row is speaker 1, leave the cell blank
        if df.iloc[0][3] == 1:
            new_row["turn " + str(turn)] = ""
            turn = 2

        while row < len(df):
            current_speaker = df.iloc[row][3]
            annotation = ""

            # Exclude data from speaker ID 3
            if row < len(df) and df.iloc[row][3] not in [1, 2]:
                row += 1
                continue

            # Merge separate lines of data from one speaker
            while row < len(df) and df.iloc[row][3] == current_speaker:
                if df.iloc[row][3] == 1 and pd.notna(df.iloc[row][5]):
                    annotation += str(df.iloc[row][5]) + ", "
                elif df.iloc[row][3] == 2:
                    annotation += str(df.iloc[row][4]) + " "
                # elif df.iloc[row][3] == 3:
                #     annotation += str(df.iloc[row][4]) + " "
                row += 1

            new_row["turn " + str(turn)] = annotation
            print(annotation)
            turn += 1

        compiled_annotations = pd.concat([compiled_annotations, pd.DataFrame([new_row])], ignore_index=True)

    compiled_annotations.to_excel(output_file, index=False)
    print("Compiled annotations saved to:", output_file)

    # Format compiled_annotations.xlsx
    workbook = openpyxl.load_workbook(output_file)
    worksheet = workbook.active

    for column in worksheet.iter_cols(min_col=4):
        # Wrap text
        for cell in column:
            cell.alignment = Alignment(wrap_text=True)

        # Resize column width
        column_letter = get_column_letter(column[0].column)
        column_width = worksheet.column_dimensions[column_letter].width
        worksheet.column_dimensions[column_letter].width = column_width * 1.5

    workbook.save(output_file)

if __name__ == '__main__':
    compileAnnotation()
    # addAddressee()

