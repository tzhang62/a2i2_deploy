import os
import pandas as pd

from openpyxl.styles import PatternFill, Alignment
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def process_csv(csv_file, columns):
    """
    Process the CSV file by selecting specific columns, rearranging columns, and adding an 'annotation' column.
    Returns the path of the modified CSV file and the processed DataFrame.
    Args:
        csv_file (str): The name of the CSV file.
        columns (list): A list of column indices to keep in the processed DataFrame.
    Returns:
        str: The path of the modified CSV file.
        pandas.DataFrame: The processed DataFrame.
    """
    csv_path = os.path.join(root, csv_file)
    file_name = os.path.splitext(os.path.basename(csv_path))[0]
    output_csv = os.path.join(root, f"{file_name}_modified.csv")
    
    df = pd.read_csv(csv_path)
    df = df.iloc[:, columns]

    last_column = df.columns[-1]
    second_to_last = df.columns[-2]
    df.insert(df.columns.get_loc(second_to_last), last_column, df.pop(last_column))

    df["annotation"] = ""

    df.to_csv(output_csv, index=False)
    print(f"Processed CSV: {csv_file}")
    return output_csv, df

def generate_html(df, output_csv):
    """
    Generate an HTML table with styling based on the values in the DataFrame.
    Saves the HTML file with the same name as the output CSV file.
    Args:
        df (pandas.DataFrame): The DataFrame containing the data.
        output_csv (str): The path of the output CSV file.
    """
    html = '<table style="border: 5px;">'
    html += '<tr>'
    for column in df.columns:
        html += f'<th style="border: 5px; background-color: gray;">{column}</th>'
    html += '</tr>'

    for row in df.values:
        if row[3] == 1: 
            row_color = 'lightgray' 
        elif row[3] == 2: 
            row_color = 'white'
        elif row[3] == 3: 
            row_color = 'darkgray'

        html += f'<tr style="background-color: {row_color};">'
        for column in row:
            html += f'<td style="border: 5px; padding-left: 5px; padding-right: 10px;">{column}</td>'
        html += '</tr>'
    html += '</table>'

    file_name = os.path.basename(os.path.splitext(output_csv)[0])
    output_html = os.path.join(os.path.dirname(output_csv), f"{file_name}.html")
    with open(output_html, 'w') as html_file:
        html_file.write(html)
    print(f"Generated HTML: {file_name}.html")

def increase_column_width(worksheet, column_index, scale):
    """
    Increase the width of a specific column in the Excel worksheet by a given scale factor.
    Args:
        worksheet: The openpyxl worksheet object.
        column_index: The index of the column to increase the width (1-based index).
        scale: The scale factor to increase the column width by.
    """
    column_letter = get_column_letter(column_index)
    column_width = worksheet.column_dimensions[column_letter].width
    worksheet.column_dimensions[column_letter].width = column_width * scale

def generate_xlsx(df, output_csv):
    """
    Generate an Excel file with formatting based on the values in the DataFrame.
    Saves the Excel file with the same name as the output CSV file.
    Args:
        df (pandas.DataFrame): The DataFrame containing the data.
        output_csv (str): The path of the output CSV file.
    """
    file_name = os.path.basename(os.path.splitext(output_csv)[0])
    output_excel = os.path.join(os.path.dirname(output_csv), f"{file_name}.xlsx")
    df.to_excel(output_excel, index=False)

    workbook = Workbook()
    sheet = workbook.active
    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)
    
    increase_column_width(sheet, 5, 3)
    increase_column_width(sheet, 6, 2)
    
    # Apply pattern fill to the header row cells
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=1, values_only=True), start=1):
        fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        for col_idx, cell in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx).fill = fill
    
    # Apply pattern fill and text wrapping to the cells in the data rows
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=len(df)+1, values_only=True), start=2): 
        if row[3] == 1: 
            fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        elif row[3] == 2: 
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        elif row[3] == 3: 
            fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

        for col_idx, cell in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx).fill = fill
            sheet.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True)

    workbook.save(output_excel)
    print(f"Generated Excel: {file_name}.xlsx")

if __name__ == '__main__':
    root_folder = "TextData"
    columns_to_keep = [1, 3, 4, 5, -1]
    
    for root, dirs, files in os.walk(root_folder):
        for file in files:
            if file.endswith("_transcription.csv"):
                output_csv, df = process_csv(file, columns_to_keep)
                generate_html(df, output_csv)
                generate_xlsx(df, output_csv)
